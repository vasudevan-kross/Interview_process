"""
Voice Screening API endpoints.

Provides REST API for:
- Creating and managing voice screening candidates
- Bulk import via CSV/Excel
- Vapi webhook handling
- Excel export
"""

import logging
import csv
import io
import json
from fastapi import APIRouter, HTTPException, status, UploadFile, File, Depends, Request
from fastapi.responses import StreamingResponse
from typing import Optional

from app.schemas.voice_screening import (
    VoiceCandidateCreate,
    VoiceCandidateBulkCreate,
)
from app.services.voice_screening_service import get_voice_screening_service
from app.auth.dependencies import get_current_user_id
from app.config import settings

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/voice-screening", tags=["voice-screening"])


@router.post("/candidates", summary="Create a voice screening candidate")
async def create_candidate(
    request: VoiceCandidateCreate,
    current_user_id: str = Depends(get_current_user_id)
):
    """Create a single voice screening candidate."""
    try:
        service = get_voice_screening_service()
        result = await service.create_candidate(
            name=request.name,
            email=request.email,
            phone=request.phone,
            is_fresher=request.is_fresher,
            user_id=current_user_id,
        )
        return result

    except Exception as e:
        logger.error(f"Error creating voice candidate: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create candidate: {str(e)}"
        )


@router.post("/candidates/bulk", summary="Bulk create voice candidates")
async def bulk_create_candidates(
    request: VoiceCandidateBulkCreate,
    current_user_id: str = Depends(get_current_user_id)
):
    """Bulk create voice screening candidates from a list."""
    try:
        service = get_voice_screening_service()
        candidates_data = [c.dict() for c in request.candidates]
        result = await service.bulk_create_candidates(candidates_data, current_user_id)
        return result

    except Exception as e:
        logger.error(f"Error bulk creating candidates: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to bulk create candidates: {str(e)}"
        )


@router.post("/candidates/upload", summary="Upload CSV/Excel of candidates")
async def upload_candidates_file(
    file: UploadFile = File(...),
    current_user_id: str = Depends(get_current_user_id)
):
    """
    Upload a CSV or Excel file to bulk create candidates.
    Expected columns: Name, Email, Phone, Is Fresher (true/false)
    """
    try:
        filename = file.filename or ""
        ext = filename.split(".")[-1].lower()

        content = await file.read()

        candidates = []

        if ext == "csv":
            # Parse CSV
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                # Normalize column names to lowercase
                row_lower = {k.strip().lower().replace(" ", "_"): v.strip() for k, v in row.items()}
                candidates.append({
                    "name": row_lower.get("name", ""),
                    "email": row_lower.get("email", ""),
                    "phone": row_lower.get("phone", "") or row_lower.get("phone_number", ""),
                    "is_fresher": row_lower.get("is_fresher", "false").lower() in ("true", "yes", "1"),
                })

        elif ext in ("xlsx", "xls"):
            # Parse Excel
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip().lower().replace(" ", "_") for cell in ws[1]]

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                candidates.append({
                    "name": str(row_dict.get("name", "") or ""),
                    "email": str(row_dict.get("email", "") or ""),
                    "phone": str(row_dict.get("phone", "") or row_dict.get("phone_number", "") or ""),
                    "is_fresher": str(row_dict.get("is_fresher", "false")).lower() in ("true", "yes", "1"),
                })
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Unsupported file type: {ext}. Use CSV or Excel (.xlsx)"
            )

        # Filter out empty rows
        candidates = [c for c in candidates if c.get("name")]

        if not candidates:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No valid candidates found in file"
            )

        service = get_voice_screening_service()
        result = await service.bulk_create_candidates(candidates, current_user_id)

        return {
            **result,
            "filename": filename,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading candidates file: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to process file: {str(e)}"
        )


@router.get("/candidates", summary="List voice screening candidates")
async def list_candidates(
    current_user_id: str = Depends(get_current_user_id),
    limit: int = 100,
    offset: int = 0,
    status_filter: Optional[str] = None
):
    """List all voice candidates created by the current user."""
    try:
        service = get_voice_screening_service()
        result = await service.list_candidates(
            user_id=current_user_id,
            limit=limit,
            offset=offset,
            status_filter=status_filter
        )
        return result

    except Exception as e:
        logger.error(f"Error listing candidates: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to list candidates: {str(e)}"
        )


@router.get("/candidates/token/{token}", summary="Get candidate by token (public)")
async def get_candidate_by_token(token: str):
    """
    Public endpoint for shareable link page.
    Returns candidate name and fresher status needed for the Vapi call config.
    """
    try:
        service = get_voice_screening_service()
        candidate = await service.get_candidate_by_token(token)

        # Only return necessary info for the public page
        return {
            "id": candidate["id"],
            "interview_token": candidate["interview_token"],
            "name": candidate["name"],
            "is_fresher": candidate["is_fresher"],
            "status": candidate["status"],
        }

    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Candidate not found"
        )
    except Exception as e:
        logger.error(f"Error getting candidate by token: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to get candidate"
        )


@router.post("/candidates/token/{token}/start-call", summary="Mark call started (public)")
async def start_call(token: str, call_id: Optional[str] = None):
    """Mark a candidate's call as in progress."""
    try:
        service = get_voice_screening_service()
        await service.update_candidate_status(token, "in_progress", call_id)
        return {"status": "in_progress"}

    except Exception as e:
        logger.error(f"Error starting call: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to start call"
        )


@router.delete("/candidates/{candidate_id}", summary="Delete a voice candidate")
async def delete_candidate(
    candidate_id: str,
    current_user_id: str = Depends(get_current_user_id)
):
    """Delete a candidate record."""
    try:
        service = get_voice_screening_service()
        return await service.delete_candidate(candidate_id, current_user_id)

    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=str(e)
        )
    except Exception as e:
        logger.error(f"Error deleting candidate: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete candidate: {str(e)}"
        )


@router.post("/webhook", summary="Vapi end-of-call webhook")
async def vapi_webhook(request: Request):
    """
    Webhook endpoint for Vapi to send end-of-call reports.
    This is called automatically by Vapi when a call ends.
    No authentication required (Vapi calls this).
    """
    try:
        body = await request.json()
        logger.info(f"Received Vapi webhook: {json.dumps(body, indent=2)[:500]}")

        message = body.get("message", {})
        msg_type = message.get("type", "")

        # Handle end-of-call report
        if msg_type == "end-of-call-report":
            call_id = message.get("call", {}).get("id", "")

            if not call_id:
                logger.warning("Webhook missing call ID")
                return {"status": "ignored", "reason": "no call_id"}

            extracted_data = {
                "transcript": message.get("transcript", ""),
                "recordingUrl": message.get("recordingUrl", ""),
                "structuredData": message.get("analysis", {}).get("structuredData", {}),
            }

            service = get_voice_screening_service()
            await service.update_candidate_from_webhook(call_id, extracted_data)

            logger.info(f"Updated candidate from call {call_id}")
            return {"status": "processed"}

        # Handle status updates
        elif msg_type == "status-update":
            call_status = message.get("status", "")
            logger.info(f"Call status update: {call_status}")
            return {"status": "acknowledged"}

        return {"status": "ignored", "type": msg_type}

    except Exception as e:
        logger.error(f"Error processing Vapi webhook: {e}")
        # Always return 200 to Vapi to prevent retries
        return {"status": "error", "message": str(e)}


@router.get("/export", summary="Export candidates to Excel")
async def export_excel(
    current_user_id: str = Depends(get_current_user_id)
):
    """Export all voice candidates to an Excel file."""
    try:
        service = get_voice_screening_service()
        excel_bytes = await service.export_to_excel(current_user_id)

        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=voice_screening_candidates.xlsx"
            }
        )

    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export: {str(e)}"
        )
